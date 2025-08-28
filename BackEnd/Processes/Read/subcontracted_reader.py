from openpyxl.utils import get_column_letter

def subcontracted_reader(ws_to_read) -> list:
    
    
    #Receive the worksheet with the subcontracted data to read
    
    registers_to_create = []
    
    columns_to_read =  [get_column_letter(i) for i in range(1, 38)]
    
    max_row = ws_to_read.max_row
    max_col = ws_to_read.max_column
    
    for row in range(1, max_row + 1):
    
        register = []
        
        
        for col in range(1, max_col + 1):
        
            
            cell_to_read = ws_to_read.cell(row = row, column= col).value

            
            register.append(cell_to_read)
        
        
        
        registers_to_create.append(register)    
     
    
    return registers_to_create