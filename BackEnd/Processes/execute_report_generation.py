import sys
import os
import subprocess
import platform
from pathlib import Path
import sys

from openpyxl import load_workbook

from BackEnd.Database.Queries.Select.select_header_data import select_header_data
from BackEnd.Database.Queries.Select.select_parameters import select_parameters
from BackEnd.Database.Queries.Select.select_quality_controls import select_quality
from BackEnd.Database.Queries.Select.select_samples import select_samples
from BackEnd.Processes.Format.block_analytical_copy import block_analitic_copy
from BackEnd.Processes.Format.block_quality_copy import block_quality_copy
from BackEnd.Processes.Format.footer_format_copy import footer_format_copy
from BackEnd.Processes.Format.header_analytic_format_copy import header_analitic_format_copy
from BackEnd.Processes.Format.header_format_copy import header_format_copy
from BackEnd.Processes.Format.header_quality_format_copy import header_quality_format_copy
from BackEnd.Processes.Format.header_summary_format_copy import header_summary_format_copy
from BackEnd.Processes.Format.lab_format_copy import lab_format_copy
from BackEnd.Processes.Format.table_final_copy import copy_table_final
from BackEnd.Processes.Read.excel_header_reader import excel_header_reader
from BackEnd.Processes.Write.write_analytic_data import write_analitic_data
from BackEnd.Processes.Write.write_lab_data import write_lab_data
from BackEnd.Processes.Write.write_quality_control import write_quality_control
from BackEnd.Utils.apply_font_to_worksheet import apply_font_to_worksheet
from BackEnd.Utils.filter_summary_data import filter_summary_data
from BackEnd.Utils.group_parameters_by_sample_id import group_parameters_by_sample
from BackEnd.Utils.pagination import pagination
from BackEnd.Utils.set_height_for_all_rows import set_height_for_all_rows





# Configuración de paths
def get_project_root():
    """Retorna el directorio raíz del proyecto"""
    return Path(__file__).parent.parent.absolute()

# Configurar paths
PROJECT_ROOT = get_project_root()
sys.path.append(str(PROJECT_ROOT))
PROJECT_DIR = Path(__file__).parent.parent.absolute()
sys.path.append(str(PROJECT_DIR))

# Definir rutas
FILE_PATH = PROJECT_ROOT / "plantilla-reporte-final.xlsx"
PATH_FILE_SOURCE = PROJECT_ROOT / "SOURCE-FORMAT.xlsx"
PATH_FILE_WRITE = PROJECT_ROOT / "Reporte.xlsx"



def open_excel_file(file_path):
    
    """
    Abre el archivo Excel usando la aplicación predeterminada del sistema
    """
    try:
        system = platform.system()
        
        if system == "Windows":
            # En Windows
            os.startfile(str(file_path))
        elif system == "Darwin":  # macOS
            # En macOS
            subprocess.run(["open", str(file_path)], check=True)
        elif system == "Linux":
            # En Linux
            subprocess.run(["xdg-open", str(file_path)], check=True)
        else:
            print(f"Sistema operativo no soportado para apertura automática: {system}")
            return False
            
        print(f"✅ Archivo Excel abierto: {file_path}")
        return True
        
    except Exception as e:
        print(f"❌ Error al abrir el archivo Excel: {str(e)}")
        return False


def execute_report_generation(batch_id, selected_samples, selected_parameters) -> bool:
        
    FILE_PATH = PROJECT_ROOT / "plantilla-reporte-final.xlsx"
    PATH_FILE_SOURCE = PROJECT_ROOT / "SOURCE-FORMAT.xlsx"
    PATH_FILE_WRITE = PROJECT_ROOT / "Reporte.xlsx"
    
    WB_TO_PRINT = load_workbook(PATH_FILE_WRITE)
    WB_TO_READ = load_workbook(filename=FILE_PATH, data_only=True)
    WB_TO_FORMAT = load_workbook(PATH_FILE_SOURCE)
    
    
    # NUEVA METODOLOGIA PARA 3 QUERYS 
    
    try:
        
        
        select_header_data(1001)
        
        
        row = 2
        last_row = pagination(WB_TO_PRINT, WB_TO_FORMAT, WB_TO_READ, row)
        header_data = excel_header_reader(WB_TO_READ)
        
        print(f"HEADER DATA {header_data}")
        client_sample_id = header_data[6]
           
        
        chain_data = select_samples(batch_id, selected_samples, None, False) 
        
        print(chain_data)
        
        
        lab_format_copy(WB_TO_FORMAT, WB_TO_PRINT, WB_TO_FORMAT["Header_lab"], last_row, len(chain_data)) 
            
             
        parameters_data = select_parameters(batch_id, selected_parameters, None)

        last_row = write_lab_data(WB_TO_PRINT, chain_data, last_row, client_sample_id)
        
        
        
        
        
        
        if last_row < 75:
            
            row_to_footer_header = last_row + (75 - last_row)
            
            footer_format_copy(WB_TO_FORMAT, WB_TO_PRINT, WB_TO_FORMAT["Footer"], row_to_footer_header)
            
            last_row = header_format_copy(WB_TO_FORMAT, WB_TO_PRINT, WB_TO_FORMAT["Header"], row_to_footer_header + 5)
            
            print(last_row + (75 - last_row))
            
            #1-73
            #74-146
            
        # !!!!!!!!!!!!!!!!!ABOUT THIS LINE IS THE ERROR THAT SAYS    EXPECTED 

        last_row_blocks = header_analitic_format_copy(WB_TO_FORMAT, WB_TO_PRINT, WB_TO_FORMAT["header_analitic"], last_row +3)
        
        print(last_row_blocks)
        
        
        
        print(parameters_data)
        
        
        
        
        samples, controls = group_parameters_by_sample(parameters_data)
        
        
        
        
        paremeters_per_block_samples = []
        samples_no_qc = []
            
        for sample, item in samples.items():
            samples_no_qc.append(sample)
            paremeters_per_block_samples.append(len(item))
        
        
        last_row = block_analitic_copy(WB_TO_FORMAT, WB_TO_PRINT, WB_TO_FORMAT["block_analitic"], last_row_blocks -1, len(samples), paremeters_per_block_samples, WB_TO_FORMAT, WB_TO_READ)
        
        
        
        
        last_row = write_analitic_data(WB_TO_FORMAT, WB_TO_PRINT, last_row_blocks, samples, last_row_blocks - 3, WB_TO_FORMAT, WB_TO_READ)
        
        
        
        last_row_blocks = header_summary_format_copy(WB_TO_FORMAT, WB_TO_PRINT, WB_TO_FORMAT["header_summary"], last_row, WB_TO_READ, WB_TO_PRINT, WB_TO_FORMAT)
        
        # Summary data 
        samples_summ = filter_summary_data(samples)
        samples_summ_no_qc = []
        parameters_per_summ = []
               
        for sample, item in samples_summ.items():
            samples_summ_no_qc.append(sample)
            parameters_per_summ.append(len(item))
               
        last_row_blocks_q = block_analitic_copy(WB_TO_FORMAT, WB_TO_PRINT, WB_TO_FORMAT["block_analitic"], last_row_blocks-2, len(samples_no_qc), parameters_per_summ, WB_TO_FORMAT, WB_TO_READ)
        
        last_row = write_analitic_data(WB_TO_FORMAT, WB_TO_PRINT, last_row_blocks, samples_summ, last_row_blocks- 4 , WB_TO_FORMAT, WB_TO_READ)
        
        
        # Quality Controls - CORRECCIÓN AQUÍ
        last_row_to_ql = header_quality_format_copy(WB_TO_FORMAT, WB_TO_PRINT, WB_TO_FORMAT["header_quality"], last_row)
            
        quality_controls = select_quality(batch_id, [])
        print(f"QUALITY CONTROLS {quality_controls}")

        # Corrección 1: Usar el worksheet directamente
        ws_report = WB_TO_PRINT["Reporte"]
        
        # Corrección 2: Llamar a block_quality_copy solo si hay controles
        

        last_row, limits_controls = block_quality_copy(WB_TO_FORMAT, WB_TO_PRINT, WB_TO_READ, last_row_to_ql-1, quality_controls)
                
    
        last_row = write_quality_control(quality_controls, last_row_to_ql, WB_TO_PRINT["Reporte"], WB_TO_FORMAT, WB_TO_PRINT, WB_TO_READ, limits_controls)
                
        
        last_row = copy_table_final(WB_TO_FORMAT, WB_TO_PRINT, "final_table", last_row + 3, WB_TO_FORMAT, WB_TO_READ)

        
       
        
        apply_font_to_worksheet(WB_TO_PRINT["Reporte"], "Calibri", 25)
        set_height_for_all_rows(WB_TO_PRINT["Reporte"], 65, 1, 5000)
        # Guardar el archivo
        WB_TO_PRINT.save(PATH_FILE_WRITE)
        #print(f"✅ Reporte generado exitosamente: {PATH_FILE_WRITE}")
        
        # Abrir automáticamente el archivo Excel
        open_excel_file(PATH_FILE_WRITE)
            
        return True
            
    except Exception as e:
        print(f"❌ Error in main_format_integrated: {str(e)}")
        return False
    

#execute_report_generation(2505017, [], [])