from pathlib import Path
import re
import sys
import os

from BackEnd.Processes.Write.write_analytic_data import write_analitic_data


# Configuración de paths
def get_project_root():
    """Retorna el directorio raíz del proyecto"""
    return Path(__file__).parent.parent.absolute()


# Configurar paths
PROJECT_ROOT = get_project_root()
sys.path.append(str(PROJECT_ROOT))
PROJECT_DIR = Path(__file__).parent.parent.absolute()
sys.path.append(str(PROJECT_DIR))

from BackEnd.Processes.Format.block_quality_copy import block_quality_copy
from BackEnd.Processes.Format.footer_format_copy import footer_format_copy
from BackEnd.Processes.Format.header_format_copy import header_format_copy
from BackEnd.Processes.Format.header_quality_format_copy import header_quality_format_copy
from BackEnd.Processes.Format.header_summary_format_copy import header_summary_format_copy
from BackEnd.Processes.Format.lab_format_copy import lab_format_copy
from BackEnd.Processes.Read.excel_header_reader import excel_header_reader
from BackEnd.Processes.Write.write_header_data import write_header_data
from BackEnd.Processes.Write.write_lab_data import write_lab_data
from BackEnd.Processes.Write.write_quality_control import write_quality_control
from BackEnd.Utils.apply_font_to_worksheet import apply_font_to_worksheet
from BackEnd.Utils.filter_summary_data import filter_summary_data
from BackEnd.Utils.group_parameters_by_sample_id import group_parameters_by_sample
from BackEnd.Utils.set_height_for_all_rows import set_height_for_all_rows

from BackEnd.Database.Queries.Select.select_parameters import select_parameters
from BackEnd.Database.Queries.Select.select_samples import select_samples
from BackEnd.Processes.Format.block_analytical_copy import block_analitic_copy
from BackEnd.Processes.Format.header_analytic_format_copy import header_analitic_format_copy

from openpyxl import load_workbook


def validate_row_number(row_value, function_name="unknown", default_value=1):
    """
    Valida que el valor de fila sea un número entero positivo
    """





    if isinstance(row_value, bool) or not isinstance(row_value, (int, float)):
        print(f"⚠️  WARNING: {function_name} returned invalid row value: {row_value}, using default: {default_value}")
        return default_value

    row_value = int(row_value)
    if row_value < 1:
        print(f"⚠️  WARNING: {function_name} returned negative row: {row_value}, using default: {default_value}")
        return default_value

    return row_value


def main_format(ln_samples: int, ln_parameters: int) -> bool:
    try:
        file_path = "C:/Users/Duban Serrano/Videos/MP/OneDrive - Laboratorio ChemiLab/Escritorio/AppSRL/plantilla-reporte-final.xlsx"
        path_file_source = "C:/Users/Duban Serrano/Videos/MP/OneDrive - Laboratorio ChemiLab/Escritorio/AppSRL/SOURCE-FORMAT.xlsx"
        path_file_write = "C:/Users/Duban Serrano/Videos/MP/OneDrive - Laboratorio ChemiLab/Escritorio/AppSRL/Reporte.xlsx"

        # Cargar workbooks
        wb_to_print = load_workbook(path_file_write)
        wb_to_read = load_workbook(filename=file_path, data_only=True)
        wb_to_format = load_workbook(path_file_source)

        # FIRST HEADER BLOCK
        print("📄 Copiando primer header...")
        last_row = header_format_copy(wb_to_format, wb_to_print, wb_to_format["Header"], 1)
        last_row = validate_row_number(last_row, "header_format_copy", 10)

        print("📖 Leyendo datos del header...")
        header_data = excel_header_reader(wb_to_read)
        write_header_data(wb_to_print, header_data, 1)
        client_sample_id = header_data[6] if len(header_data) > 6 else None

        # Lab data block
        print("🧪 Copiando formato de laboratorio...")
        lab_result = lab_format_copy(wb_to_format, wb_to_print, wb_to_format["Header_lab"], last_row, 48)
        if lab_result is not None:
            last_row = validate_row_number(lab_result, "lab_format_copy", last_row + 10)

        print("📊 Obteniendo datos de muestras y parámetros...")
        chain_data = select_samples(2504020, [], None, False)
        parameters_data = select_parameters(2504020, [])

        print("✍️  Escribiendo datos de laboratorio...")
        lab_write_result = write_lab_data(wb_to_print, chain_data, last_row, client_sample_id)
        last_row = validate_row_number(lab_write_result, "write_lab_data", last_row + 5)

        # Footer y segundo header
        print("🦶 Copiando footer...")
        footer_result = footer_format_copy(wb_to_format, wb_to_print, wb_to_format["Footer"], last_row)
        last_row_to_header = validate_row_number(footer_result, "footer_format_copy", last_row + 5)

        print("📄 Copiando segundo header...")
        header2_result = header_format_copy(wb_to_format, wb_to_print, wb_to_format["Header"], last_row_to_header)
        last_row = validate_row_number(header2_result, "header_format_copy (second)", last_row_to_header + 10)

        write_header_data(wb_to_print, header_data, last_row_to_header)

        # Header analítico
        print("🔬 Copiando header analítico...")
        analytic_header_result = header_analitic_format_copy(wb_to_format, wb_to_print, wb_to_format["header_analitic"],
                                                             last_row)
        last_row_blocks = validate_row_number(analytic_header_result, "header_analitic_format_copy", last_row + 5)

        times_to_print_blocks = len(chain_data)

        print("🔄 Agrupando parámetros por muestra...")
        try:
            samples, controls = group_parameters_by_sample(parameters_data)
        except Exception as e:
            print(f"❌ Error al agrupar parámetros: {e}")
            print("ℹ️  Verificar la función group_parameters_by_sample")
            return False

        paremeters_per_block_samples = []
        samples_no_qc = []

        for sample, item in samples.items():
            samples_no_qc.append(sample)
            paremeters_per_block_samples.append(len(item))

        print("📋 Copiando bloques analíticos...")
        block_result = block_analitic_copy(wb_to_format, wb_to_print, wb_to_format["block_analitic"],
                                           last_row_blocks, len(samples_no_qc), paremeters_per_block_samples)
        last_row = validate_row_number(block_result, "block_analitic_copy", last_row_blocks + 10)

        print("✍️  Escribiendo datos analíticos...")
        try:

            print(f"DEBUG: Calling write_analitic_data with:")
            print(f"  wb_to_format: {type(wb_to_format)}")
            print(f"  wb_to_print: {type(wb_to_print)}")
            print(f"  last_row_blocks: {last_row_blocks}")
            print(f"  samples count: {len(samples)}")



            analytic_write_result = write_analitic_data(wb_to_format, wb_to_print, last_row_blocks, samples,
                                                        last_row_blocks)


            print(f"RESULTADO DE ANALYTIC WRITE RESULT {analytic_write_result}")


            last_row = validate_row_number(analytic_write_result, "write_analitic_data", last_row + 5)
        except Exception as e:
            print(f"❌ ERROR al escribir datos analíticos: {e}")
            last_row = last_row + 10  # Valor por defecto si falla

        # Summary data
        print("📝 Procesando datos de resumen...")
        summary_header_result = header_summary_format_copy(wb_to_format, wb_to_print, wb_to_format["header_summary"],
                                                           last_row)
        last_row_blocks = validate_row_number(summary_header_result, "header_summary_format_copy", last_row + 5)

        samples_summ = filter_summary_data(samples)
        samples_summ_no_qc = []
        parameters_per_summ = []

        for sample, item in samples_summ.items():
            samples_summ_no_qc.append(sample)
            parameters_per_summ.append(len(item))

        summary_block_result = block_analitic_copy(wb_to_format, wb_to_print, wb_to_format["block_analitic"],
                                                   last_row_blocks, len(samples_no_qc), parameters_per_summ)
        last_row = validate_row_number(summary_block_result, "block_analitic_copy (summary)", last_row_blocks + 5)

        summary_write_result = write_analitic_data(wb_to_format, wb_to_print, last_row_blocks, samples_summ,
                                                   last_row_blocks)
        last_row = validate_row_number(summary_write_result, "write_analitic_data (summary)", last_row + 5)

        # Quality Controls
        print("🎯 Procesando controles de calidad...")
        quality_header_result = header_quality_format_copy(wb_to_format, wb_to_print, wb_to_format["header_quality"],
                                                           last_row)
        last_row = validate_row_number(quality_header_result, "header_quality_format_copy", last_row + 5)

        # Validar que last_row - 1 sea positivo
        quality_start_row = max(1, last_row - 1)
        block_quality_copy(wb_to_format, wb_to_print, quality_start_row, controls)

        def transform_data(input_data):

            """
            Transforma la estructura de datos original en el formato deseado.
            Extrae las primeras 15 posiciones de cada lista y sublista.
            """

            result = {}

            for key, value in input_data.items():
                result[key] = []
                result[key].append(value[:15])

                for item in value:
                    if isinstance(item, list):
                        result[key].append(item[:15])

            return result

        print("🔄 Transformando datos de control...")
        controls_org = transform_data(controls)

        quality_write_start_row = max(1, last_row - 2)
        write_quality_control(controls_org, quality_write_start_row, wb_to_print["Reporte"])

        print("🎨 Aplicando formato final...")
        set_height_for_all_rows(wb_to_print["Reporte"], 60, 1, 5000)
        apply_font_to_worksheet(wb_to_print["Reporte"], "Calibri", 25)

        print("💾 Guardando archivo...")
        wb_to_print.save(path_file_write)

        print("✅ Proceso completado exitosamente!")
        return True

    except Exception as e:
        print(f"❌ Error in main_format_integrated: {e}")
        import traceback
        traceback.print_exc()
        return False