

from BackEnd.Utils.copy_excel_range import copy_excel_range
from BackEnd.Utils.pagination import check_pagination_needed, pagination


def block_quality_copy(wb_format, wb_to_print, wb_to_read, last_row: int, controls):
    #print("\n=== DEBUG: STARTING block_quality_copy ===")
    #print(f"Initial last_row: {last_row}")
    #print(f"Number of controls to process: {len(controls)}")
    #print(f"Source workbook sheets: {wb_format.sheetnames}")
    #print(f"Destination workbook sheets: {wb_to_print.sheetnames}")
    
    # Lista para almacenar los rangos de filas de cada bloque
    block_ranges = []
    
    # Verify destination sheet exists
    if "Reporte" not in wb_to_print.sheetnames:
        #print("!!! CRITICAL ERROR: 'Reporte' sheet not found in destination workbook !!!")
        return last_row, block_ranges
    
    # NO USAR pagination() AQUÍ - solo procesar los controles individualmente
    # La paginación inicial ya se maneja en el flujo principal
    
    for i, control in enumerate(controls):
       
        
        # Verificar que el control es una lista válida y tiene al menos 14 elementos (índice 13)
        if not isinstance(control, list) or len(control) < 14:
            #print(f"!!! WARNING: Control is not a valid list or doesn't have enough elements (length: {len(control) if isinstance(control, list) else 'N/A'}) !!!")
            continue
        
        # Obtener el tipo de control de la posición 13 (índice 13, no 12)
        control_type = control[13]
        
        if control_type is None:
            #print("!!! WARNING: Control type is None !!!")
            continue
        
        try:
            # Mapear el control a su hoja correspondiente
            sheet_name = map_control_to_sheet(control_type)
            
            # Verificar que la hoja existe
            if sheet_name not in wb_format.sheetnames:
                #print(f"!!! ERROR: HOJA NO ENCONTRADA: {sheet_name} !!!")
                #print(f"Hojas disponibles: {wb_format.sheetnames}")
                continue
            
            sheet_to_copy = wb_format[sheet_name]
            #print(f"Found source sheet: {sheet_name} with max_row: {sheet_to_copy.max_row}")
            
            # USAR LA FUNCIÓN PAGINATION DEDICADA CON LA LÓGICA CORRECTA
            rows_needed = 8  # Cada bloque de formato necesita 8 filas
            #print(f"Checking pagination: current_row={last_row}, rows_needed={rows_needed}")
            
            # Verificar si necesitamos paginación usando la función correcta
            needs_pagination, page_end, available_rows = check_pagination_needed(last_row, rows_needed)
            
            if needs_pagination:
                #print(f"PAGINATION NEEDED: current row {last_row}, needs {rows_needed} rows, available {available_rows}")
                
                # USAR LA FUNCIÓN PAGINATION DEDICADA - CORRECCIÓN AQUÍ
                last_row = pagination(wb_to_print, wb_format, wb_to_read, last_row, rows_needed)
                #print(f"After pagination, new last_row: {last_row}")
            
            # Guardar la fila inicial del bloque actual
            block_start_row = last_row
            
            range_to_copy = "A1:AQ8"  # Ajusta este rango según tu formato
            
            # Debug source range
            try:
                from openpyxl.utils import range_boundaries
                src_start_col, src_start_row, src_end_col, src_end_row = range_boundaries(range_to_copy)
                #print(f"Source range boundaries: cols {src_start_col}-{src_end_col}, rows {src_start_row}-{src_end_row}")
            except Exception as e:
                print(f"!!! ERROR parsing source range: {e} !!!")
            
            # Perform the copy
            success = copy_excel_range(sheet_to_copy, wb_to_print["Reporte"], range_to_copy, f"A{last_row}")
            
            if not success:
                print("!!! COPY OPERATION FAILED !!!")
                continue
            
            print(f"Successfully copied format to row {last_row}")
            
            # Avanzar 8 filas para el siguiente formato
            last_row += 8
            
            # Calcular la fila final del bloque (última fila ocupada)
            block_end_row = last_row - 1
            
            # Agregar el rango del bloque a la lista
            block_ranges.append([block_start_row, block_end_row])
            #print(f"Block range added: [{block_start_row}, {block_end_row}]")
            #print(f"Updated last_row: {last_row}")
                
        except Exception as e:
            #print(f"!!! EXCEPTION processing control: {e} !!!")
            import traceback
            traceback.print_exc()
            continue
    
    #print(f"\n=== DEBUG: FINISHED block_quality_copy ===")
    #print(f"Final last_row: {last_row}")
    #print(f"Block ranges: {block_ranges}")
    return last_row, block_ranges

def map_control_to_sheet(control_type):
    #print(f"\n[DEBUG map_control_to_sheet] Input: {control_type} (type: {type(control_type)})")
    
    if control_type is None:
        #print("[DEBUG] Control type is None, returning 'unknown'")
        return 'unknown'
        
    # Convertir a string y normalizar
    control_type_str = str(control_type).strip().upper()
    #print(f"[DEBUG] Normalized control type: '{control_type_str}'")
    
    # Mapeo directo basado en los tipos que mencionas
    type_mapping = {
        'DUP': 'DUP',
        'LCS': 'LCS', 
        'LCSD': 'LCSD',
        'MB': 'MB',
        'MS': 'MS',
        'MSD': 'MSD'
    }
    
    # Buscar coincidencia exacta
    if control_type_str in type_mapping:
        sheet_name = type_mapping[control_type_str]
        #print(f"[DEBUG] Direct match found: {sheet_name}")
        return sheet_name
    
    # Si no hay coincidencia exacta, buscar como substring
    for key, value in type_mapping.items():
        if key in control_type_str:
            #print(f"[DEBUG] Substring match found: {key} -> {value}")
            return value
    
    # Si no se encuentra ninguna coincidencia
    print(f"[DEBUG] No match found for '{control_type_str}', returning 'unknown'")
    return 'unknown'