from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.cell.cell import MergedCell
import re  # Agrega la importación de re


def merged_cell(ws, celda_coord, valor) -> bool:
    """
    Escribe un valor en una celda específica en una hoja de Excel, manejando celdas combinadas.

    Esta función escribe un valor en una coordenada de celda específica en una hoja de Excel.
    Cuando la celda objetivo es parte de un rango combinado, la función identifica la celda principal
    (celda superior izquierda) del rango combinado y escribe el valor allí, manteniendo el comportamiento
    correcto de Excel.

    Args:
        ws: El objeto hoja de trabajo (openpyxl Worksheet)
        celda_coord (str): Coordenada de celda en formato Excel (ej. 'A1', 'B12')
        valor: El valor a escribir en la celda

    Returns:
        bool: True si la operación de escritura fue exitosa, False en caso contrario
    """
    try:
        # Analizar la coordenada de la celda usando regex para extraer letras de columna y número de fila
        match = re.match(r'([A-Za-z]+)(\d+)', celda_coord)
        if not match:
            print(f"Formato de coordenada incorrecto: {celda_coord}")
            return False

        # Extraer letras de columna y número de fila del match
        col_str, row_str = match.groups()
        row = int(row_str)
        col = column_index_from_string(col_str)

        # Verificar si la celda está en un rango combinado
        for rango in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = rango.min_row, rango.min_col, rango.max_row, rango.max_col

            # Verificar si la celda objetivo está dentro de este rango combinado
            if min_row <= row <= max_row and min_col <= col <= max_col:
                # Obtener la celda principal (superior izquierda) del rango combinado
                celda_principal = ws.cell(row=min_row, column=min_col)
                # Escribir el valor en la celda principal
                celda_principal.value = valor
                """print(
                    f"Escribiendo '{valor}' en la celda principal {get_column_letter(min_col)}{min_row} del rango combinado")"""
                return True

        # Si no está en un rango combinado, escribir directamente
        # NOTA: No es necesario verificar si es una MergedCell aquí,
        # ya que ya hemos verificado todos los rangos combinados
        ws.cell(row=row, column=col).value = valor
        #print(f"Escribiendo '{valor}' directamente en {celda_coord}")
        return True

    except Exception as e:
        # Capturar cualquier error inesperado que pueda ocurrir
        #print(f"Error en write_cell con coordenada {celda_coord}: {str(e)}")
        return False


