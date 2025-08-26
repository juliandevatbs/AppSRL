import re
import math
from openpyxl.cell import MergedCell
from openpyxl.utils import column_index_from_string

def write_cell(ws, celda_coord, valor) -> bool:
    """
    Write a value to a specified cell in an Excel worksheet, handling merged cells.
     
    This function writes a value to a specified cell coordinate in an Excel worksheet.
    When the target cell is part of a merged range, the function identifies the primary cell
    (top-left cell) of the merged range and writes the value there, maintaining proper Excel behavior.
    If the value is None, empty string, whitespace only, or any invalid value, it writes "-" instead.
     
    Args:
        ws: The worksheet object (openpyxl Worksheet)
        celda_coord (str): Cell coordinate in Excel format (e.g., 'A1', 'B12')
        valor: The value to write to the cell
     
    Returns:
        bool: True if the write operation was successful, False otherwise
    """
    try:
        # Función auxiliar para verificar si un valor es inválido
        def is_invalid_value(val):
            """
            Check if a value should be considered invalid and replaced with "-"
            """
            # None values
            if val is None:
                return True
            
            # String checks
            if isinstance(val, str):
                # Empty string or whitespace only
                if val.strip() == "":
                    return True
                # Common representations of null/empty values
                if val.lower().strip() in ['null', 'none', 'n/a', 'na', '#n/a', 'nil', '']:
                    return True
            
            # Numeric checks
            if isinstance(val, (int, float)):
                # NaN (Not a Number)
                if math.isnan(val):
                    return True
                # Infinite values
                if math.isinf(val):
                    return True
            
            # Boolean False could be valid, so we don't treat it as invalid
            # Empty collections (lists, tuples, dicts) - treat as invalid
            if isinstance(val, (list, tuple, dict, set)) and len(val) == 0:
                return True
                
            return False
        
        # Check if the value is invalid and replace with "-"
        if is_invalid_value(valor):
            valor = "-"
        
        # Parse the cell coordinate using regex to extract column letters and row number
        match = re.match(r'([A-Za-z]+)(\d+)', celda_coord)
        if not match:
            print(f"Incorrect coords: {celda_coord}")
            return False
         
        # Extract column letters and row number from the match
        col_str, row_str = match.groups()
        row = int(row_str)
        col = column_index_from_string(col_str)
         
        # Get the cell at the specified coordinate
        celda = ws.cell(row=row, column=col)
         
        # If it's a regular cell (not merged), write the value directly
        if not isinstance(celda, MergedCell):
            celda.value = valor
            return True
         
        # If it's a merged cell, find the primary cell in the merged range
        for rango in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = rango.min_row, rango.min_col, rango.max_row, rango.max_col
             
            # Check if the target cell is within this merged range
            if min_row <= row <= max_row and min_col <= col <= max_col:
                # Get the primary cell (top-left) of the merged range
                celda_principal = ws.cell(row=min_row, column=min_col)
                # Write the value to the primary cell
                celda_principal.value = valor
                return True
         
        # If no matching merged range was found (unusual case)
        print(f"Not found {celda_coord}")
        return False
     
    except Exception as e:
        # Catch any unexpected errors that might occur
        print(f"Error {celda_coord}: {str(e)}")
        return False

"""# Ejemplos de uso y casos de prueba
if __name__ == "__main__":
    # Casos que deberían ser reemplazados por "-":
    test_cases = [
        None,
        "",
        "   ",  # Solo espacios
        "null",
        "NULL",
        "None",
        "NONE",
        "n/a",
        "N/A",
        "#N/A",
        "na",
        "NA",
        "nil",
        "NIL",
        float('nan'),  # NaN
        float('inf'),  # Infinito
        float('-inf'), # Infinito negativo
        [],  # Lista vacía
        {},  # Diccionario vacío
        (),  # Tupla vacía
        set(),  # Set vacío
    ]
    
    # Casos que NO deberían ser reemplazados:
    valid_cases = [
        0,  # Cero es válido
        False,  # False booleano es válido
        "0",  # String "0" es válido
        "false",  # String "false" es válido
        -1,  # Números negativos son válidos
        1.5,  # Números decimales son válidos
        "texto válido",
        [1, 2, 3],  # Lista con elementos
        {"key": "value"},  # Diccionario con elementos
    ]
    
    print("=== CASOS QUE DEBERÍAN SER REEMPLAZADOS POR '-' ===")
    for case in test_cases:
        # Simular la función is_invalid_value
        def is_invalid_value(val):
            if val is None:
                return True
            if isinstance(val, str):
                if val.strip() == "":
                    return True
                if val.lower().strip() in ['null', 'none', 'n/a', 'na', '#n/a', 'nil', '']:
                    return True
            if isinstance(val, (int, float)):
                if math.isnan(val):
                    return True
                if math.isinf(val):
                    return True
            if isinstance(val, (list, tuple, dict, set)) and len(val) == 0:
                return True
            return False
        
        result = "-" if is_invalid_value(case) else case
        print(f"{repr(case)} -> {repr(result)}")
    
    print("\n=== CASOS QUE NO DEBERÍAN SER REEMPLAZADOS ===")
    for case in valid_cases:
        def is_invalid_value(val):
            if val is None:
                return True
            if isinstance(val, str):
                if val.strip() == "":
                    return True
                if val.lower().strip() in ['null', 'none', 'n/a', 'na', '#n/a', 'nil', '']:
                    return True
            if isinstance(val, (int, float)):
                if math.isnan(val):
                    return True
                if math.isinf(val):
                    return True
            if isinstance(val, (list, tuple, dict, set)) and len(val) == 0:
                return True
            return False
        
        result = "-" if is_invalid_value(case) else case
        print(f"{repr(case)} -> {repr(result)}")"""